import openpyxl

# Define the paths to your text file and Excel file
text_file_path = r"C:\Users\shaun\Desktop\Marketplace_Scripting\amazon_links.txt"
excel_file_path = r"C:\Users\shaun\Desktop\Marketplace_Scripting\FB_marketplace_2023_Grind.xlsx"

# Open the text file and read the links
with open(text_file_path, "r") as file:
    links = file.readlines()

# Open the Excel file
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook.active

# Find the next empty row in column A
next_row = sheet.max_row + 1

# Add the links to column A
for link in links:
    sheet.cell(row=next_row, column=1, value=link.strip())
    next_row += 1

# Save the modified Excel file
workbook.save(excel_file_path)

print("Links added to the Excel file.")
